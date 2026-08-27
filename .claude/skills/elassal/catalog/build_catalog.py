#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ElAssal catalog builder — maps client SKUs onto the ITEM TREE and emits
the five deliverables. The client's SKUs.xlsx is read-only / never modified."""
import json, re, os, unicodedata
from collections import Counter, defaultdict, OrderedDict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
SKUS = os.path.join(ROOT, 'data', 'SKUs.xlsx')
TREE = os.path.join(ROOT, 'data', 'ITEM TREE.xlsx')

# ---------- Arabic normalization ----------
_TASHKEEL = re.compile(r'[ؗ-ًؚ-ْٰـ]')  # harakat + tatweel
_AR2WEST = {ord(c): str(i) for i, c in enumerate('٠١٢٣٤٥٦٧٨٩')}
_AR2WEST.update({ord(c): str(i) for i, c in enumerate('۰۱۲۳۴۵۶۷۸۹')})
def norm(s):
    if s is None: return ''
    s = str(s).strip()
    s = s.translate(_AR2WEST)
    s = _TASHKEEL.sub('', s)
    s = (s.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
           .replace('ة', 'ه').replace('ى', 'ي').replace('ؤ', 'و').replace('ئ', 'ي'))
    s = s.replace('/', ' ').replace('\\', ' ')
    s = re.sub(r'\s+', ' ', s).strip().lower()
    return s
def tokens(s):
    return [t for t in norm(s).split(' ') if t]

def slugify(name):
    s = str(name).replace('/', ' ').strip()
    s = _TASHKEEL.sub('', s)
    s = re.sub(r'\s+', '-', s)
    s = re.sub(r'[^\w؀-ۿ\-]', '', s)
    return s.strip('-') or 'cat'

# ---------- Brand synonyms (for combined brand+product search) ----------
BRAND_SYNONYMS = {
    'فيات':       ['fiat', 'fiatagri', 'فيات اجري', 'fiat agri'],
    'بركنز':      ['perkins', 'بيركنز', 'بركينز'],
    'جوندير':     ['john deere', 'johndeere', 'جون دير', 'دير'],
    'فورد':       ['ford', 'فورد تراكتور'],
    'دويتس':      ['deutz', 'دوتز'],
    'نصر':        ['nasr'],
    'روسي':       ['rossi', 'russian', 'روسى'],
    'كمنز':       ['cummins', 'كومنز'],
    'روماني':     ['romani', 'romanian', 'رومانى'],
    'زيتور':      ['zetor', 'زتور'],
    'ماجروس':     ['magirus', 'ماغيروس'],
    'نيوهولند':   ['new holland', 'newholland', 'نيو هولند'],
    'ماسي':       ['massey', 'massey ferguson', 'ماسي فيرجسون', 'فيرجسون'],
    'سام':        ['same', 'سامي'],
    'لمبرجيني':   ['lamborghini'],
    'ايفيكو':     ['iveco'],
    'جي سي بي':   ['jcb'],
    'يوتب':       ['utb'],
}
def brand_syn(b):
    nb = norm(b)
    for ar, syns in BRAND_SYNONYMS.items():
        if norm(ar) == nb or nb in [norm(x) for x in syns]:
            return [ar] + syns
    return [b]

# ---------- Alias table: unmatched product item-types -> tree leaf ----------
ALIAS = {
    'عمود كرنك محرك':       'عمود كرنك',
    'جلب كامة / كرنك':      'جلبة كرنك',
    'اويل سيل كرنك خلفي':   'اويل سيل كرنك',
    'اويل سيل كرنك امامي':  'اويل سيل كرنك',
    'ريش ديسك دبرياج':      'ريشة ديسك دبرياج',
    'كارجة جاز كهرباء':     'كارجة جاز',
    'عمود / بلية طلمبة مياة':'عمود طلمبة مياة',
    'بلاطة ديسك':           'بلاطة ديسك دبرياج',
    'غطاء ردياتير مياة':    'غطاء ردياتير',
    'مسمار حدافة':          'مسمار حدافة / فلام',
    'وش ديسك':              'ديسك دبرياج',
    'وصلة طلمبة مياة':      'طلمبة مياة',
    'قاعدة ردياتير مياة':   'قاعدة ردياتير',
    'بلف طلمبة زيت':        'بلف زيت',
    'وصلة طلمبة زيت':       'طلمبة زيت محرك',
    'كوعة طلمبة هيدروليك':  'طلمبة هيدروليك / دريكسيون / باور',
    'بلف كارجة جاز':        'بلف كارجة',
}

# ---------- Load ITEM TREE ----------
wb = openpyxl.load_workbook(TREE, read_only=True, data_only=True)
rows = list(wb['Sheet1'].iter_rows(values_only=True))[1:]
parent_of, level_of, children = {}, {}, defaultdict(list)
order = []
for cat, parent, level, main in rows:
    if cat is None: continue
    c = str(cat).strip(); p = str(parent).strip() if parent else None
    if not c: continue  # skip empty-named rows
    if c in parent_of: continue  # skip duplicate rows
    parent_of[c] = p; order.append(c)
    children[p].append(c)
# compute level from parent depth (the Level column has nulls)
def _depth(c):
    d = 1; cur = parent_of.get(c)
    while cur is not None:
        d += 1; cur = parent_of.get(cur)
    return d
for c in order:
    level_of[c] = _depth(c)
leaves = set(c for c in parent_of if c not in children)
norm_name_to_node = {}
for c in parent_of:
    norm_name_to_node.setdefault(norm(c), c)

def chain(node):
    """return (l1,l2,l3) walking up to roots"""
    path = []
    cur = node
    while cur is not None:
        path.append(cur)
        cur = parent_of.get(cur)
    path = path[::-1]  # root..node
    l1 = path[0] if len(path) >= 1 else None
    l2 = path[1] if len(path) >= 2 else None
    l3 = path[2] if len(path) >= 3 else None
    # if matched node deeper than 3 just keep first three; if shallower l3/l2 may be None
    return l1, l2, l3, level_of.get(node)

# ---------- Load products ----------
wb2 = openpyxl.load_workbook(SKUS, read_only=True, data_only=True)
prod_rows = list(wb2['المنتجات'].iter_rows(values_only=True))[2:]

# candidate nodes for name inference: L2 + L3 (skip broad L1 roots).
# sort longest-first; leaves win ties so we place as deep/specific as possible.
infer_cands = [(norm(c), c) for c in parent_of if level_of[c] >= 2]
infer_cands.sort(key=lambda x: (len(x[0]), 1 if x[1] in leaves else 0), reverse=True)

def infer_from_name(name):
    nn = norm(name)
    # prefer node whose normalized name is a prefix token of the product name
    for ln, node in infer_cands:
        if not ln: continue
        if nn == ln or nn.startswith(ln + ' '):
            return node
    # fallback: appears as a standalone token sequence anywhere in the name
    for ln, node in infer_cands:
        if not ln: continue
        if (' ' + ln + ' ') in (' ' + nn + ' '):
            return node
    return None

products = []
method_counts = Counter()
cat_counts = Counter()        # product count per node (leaf/assigned)
brand_counts = Counter()
brand_models = defaultdict(set)
l1_brand = defaultdict(Counter)   # main-system -> brand -> count
uncategorized = []

for r in prod_rows:
    if r[0] is None: continue
    sku = r[0]
    name = (str(r[1]).strip() if r[1] else '')
    itype = (str(r[2]).strip() if r[2] else '')
    brand = (str(r[3]).strip() if r[3] else '')
    model = (str(r[4]).strip() if r[4] else '')
    matched = None; method = None
    if itype:
        if itype in leaves:
            matched, method = itype, 'exact-leaf'
        elif itype in parent_of:
            matched, method = itype, 'node'
        elif itype in ALIAS:
            matched, method = ALIAS[itype], 'alias'
        else:
            ni = norm_name_to_node.get(norm(itype))
            if ni:
                matched, method = ni, 'exact-leaf' if ni in leaves else 'node'
            else:
                matched = infer_from_name(itype + ' ' + name)
                method = 'name-inferred' if matched else None
    else:
        matched = infer_from_name(name)
        method = 'name-inferred' if matched else None
    if matched is None:
        method = 'uncategorized'
        l1 = l2 = l3 = None; mlevel = None
    else:
        l1, l2, l3, mlevel = chain(matched)
        cat_counts[matched] += 1
    method_counts[method] += 1
    if brand:
        brand_counts[brand] += 1
        if model: brand_models[brand].add(model)
        if l1: l1_brand[l1][brand] += 1   # brand counts within each main system
    if method == 'uncategorized':
        uncategorized.append({'sku': sku, 'name': name, 'item_type': itype, 'brand': brand})

    # search blob: name + type + brand(+synonyms) + model
    blob_parts = [name, itype, matched or '', model]
    for s in brand_syn(brand):
        blob_parts.append(s)
    search_tokens = sorted(set(t for p in blob_parts for t in tokens(p)))

    products.append(OrderedDict([
        ('sku', sku), ('name', name), ('brand', brand), ('model', model),
        ('item_type', itype),
        ('l1', l1), ('l2', l2), ('l3', l3),
        ('matched_node', matched), ('match_method', method),
        ('search', ' '.join(search_tokens)),
    ]))

# ---------- categories.json ----------
used_slugs = {}
def uniq_slug(name, nid):
    base = slugify(name)
    if base not in used_slugs:
        used_slugs[base] = nid; return base
    return base + '-' + str(nid)

cat_list = []
node_id = {}
for i, c in enumerate(order, 1):
    node_id[c] = i
for c in order:
    nid = node_id[c]
    p = parent_of.get(c)
    cat_list.append(OrderedDict([
        ('id', nid),
        ('name_ar', c),
        ('slug', uniq_slug(c, nid)),
        ('level', level_of[c]),
        ('parent_id', node_id.get(p) if p else None),
        ('is_leaf', c in leaves),
        ('product_count', cat_counts.get(c, 0)),
    ]))
# roll up counts to ancestors — count EVERY node's own products (a product
# may be assigned to a non-leaf L2 node, e.g. filters), not just leaves.
id_to_node = {n['id']: n for n in cat_list}
total_by_id = defaultdict(int)
for n in cat_list:
    if n['product_count']:
        total_by_id[n['id']] += n['product_count']
        pid = n['parent_id']
        while pid is not None:
            total_by_id[pid] += n['product_count']
            pid = id_to_node[pid]['parent_id']
for n in cat_list:
    n['product_count_total'] = total_by_id[n['id']]

os.makedirs(HERE, exist_ok=True)
with open(os.path.join(HERE, 'categories.json'), 'w', encoding='utf-8') as f:
    json.dump(cat_list, f, ensure_ascii=False, indent=2)
with open(os.path.join(HERE, 'products.json'), 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

# ---------- search-index.json ----------
search_index = OrderedDict([
    ('normalization', {
        'arabic_indic_digits': 'mapped to western',
        'tatweel_and_tashkeel': 'stripped',
        'alef_variants': 'أإآ -> ا', 'taa_marbuta': 'ة -> ه',
        'alef_maqsura': 'ى -> ي', 'hamza_seats': 'ؤ->و ئ->ي', 'slashes': 'treated as space',
    }),
    ('match_rule', 'tokenize query; AND-match all tokens against product.search (order-independent). Brand+part combos resolve via brand synonyms.'),
    ('brand_synonyms', {ar: [ar] + syns for ar, syns in BRAND_SYNONYMS.items()}),
    ('products', [OrderedDict([('sku', p['sku']), ('search', p['search']),
                               ('brand', p['brand']), ('l1', p['l1']), ('l3', p['l3'])]) for p in products]),
])
with open(os.path.join(HERE, 'search-index.json'), 'w', encoding='utf-8') as f:
    json.dump(search_index, f, ensure_ascii=False, indent=2)

# ---------- navbar.json (dual-path, grouped) ----------
TOP_L3 = 6
roots = [n for n in cat_list if n['level'] == 1]
roots.sort(key=lambda n: -n['product_count_total'])
shop_by_part = []
for r in roots:
    l2s = [n for n in cat_list if n['parent_id'] == r['id']]
    l2s.sort(key=lambda n: -n['product_count_total'])
    cols = []
    for g in l2s:
        if g['product_count_total'] == 0: continue
        l3s = [n for n in cat_list if n['parent_id'] == g['id'] and n['product_count_total'] > 0]
        l3s.sort(key=lambda n: -n['product_count_total'])
        cols.append(OrderedDict([
            ('title', g['name_ar']), ('slug', g['slug']),
            ('count', g['product_count_total']),
            ('items', [OrderedDict([('name', x['name_ar']), ('slug', x['slug']), ('count', x['product_count_total'])]) for x in l3s[:TOP_L3]]),
            ('view_all', len(l3s) > TOP_L3),
            ('total_subitems', len(l3s)),
        ]))
    sys_brands = [OrderedDict([('name', b), ('count', c)])
                  for b, c in l1_brand[r['name_ar']].most_common()]
    shop_by_part.append(OrderedDict([
        ('title', r['name_ar']), ('slug', r['slug']),
        ('count', r['product_count_total']), ('columns', cols),
        ('brands', sys_brands),
    ]))

shop_by_brand = [OrderedDict([('name', b), ('count', c), ('slug', slugify(b)),
                             ('synonyms', brand_syn(b))])
                 for b, c in brand_counts.most_common()]
shop_by_machine = [OrderedDict([('brand', b), ('count', brand_counts[b]),
                               ('models', sorted(brand_models[b]))])
                   for b, c in brand_counts.most_common() if brand_models[b]]

navbar = OrderedDict([
    ('primary', [
        OrderedDict([('label', 'تسوق حسب القطعة'), ('type', 'mega'), ('groups', shop_by_part)]),
        OrderedDict([('label', 'تسوق حسب الماركة'), ('type', 'brands'), ('items', shop_by_brand)]),
        OrderedDict([('label', 'تسوق حسب الماكينة'), ('type', 'machines'), ('items', shop_by_machine)]),
        OrderedDict([('label', 'العروض'), ('type', 'link'), ('href', '/offers')]),
    ]),
    ('support', [
        {'label': 'واتساب', 'type': 'whatsapp', 'value': '+201223149614'},
        {'label': 'الخط الساخن', 'type': 'phone', 'value': '+201223149614'},
        {'label': 'الفروع', 'type': 'branches', 'href': '/branches'},
    ]),
    ('funnel', ['PLP', 'PDP', 'Cart', 'Checkout', 'OrderSuccess']),
])
with open(os.path.join(HERE, 'navbar.json'), 'w', encoding='utf-8') as f:
    json.dump(navbar, f, ensure_ascii=False, indent=2)

# ---------- mapping-report.md ----------
total = len(products)
mapped = total - method_counts['uncategorized']
lines = []
lines.append('# ElAssal Catalog — Mapping Report')
lines.append('')
lines.append(f'- **Total products:** {total}')
lines.append(f'- **Mapped to a category:** {mapped} ({mapped*100//total}%)')
lines.append(f'- **Uncategorized (needs client review):** {method_counts["uncategorized"]} ({method_counts["uncategorized"]*100//total}%)')
lines.append('')
lines.append('## Coverage by match method')
lines.append('| Method | Products |')
lines.append('|---|---|')
for m in ['exact-leaf', 'alias', 'node', 'name-inferred', 'uncategorized']:
    lines.append(f'| {m} | {method_counts.get(m,0)} |')
lines.append('')
lines.append(f'## Brands ({len(brand_counts)})')
lines.append('| Brand | Products | Models |')
lines.append('|---|---|---|')
for b, c in brand_counts.most_common():
    lines.append(f'| {b} | {c} | {len(brand_models[b])} |')
lines.append('')
lines.append('## Main categories (L1) by product count')
lines.append('| Category | Products |')
lines.append('|---|---|')
for r in roots:
    lines.append(f'| {r["name_ar"]} | {r["product_count_total"]} |')
lines.append('')
lines.append(f'## Uncategorized list ({len(uncategorized)})')
if uncategorized:
    lines.append('| SKU | Name | item_type | Brand |')
    lines.append('|---|---|---|---|')
    for u in uncategorized:
        lines.append(f'| {u["sku"]} | {u["name"]} | {u["item_type"]} | {u["brand"]} |')
else:
    lines.append('_None — every product mapped._')
with open(os.path.join(HERE, 'mapping-report.md'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print('DONE')
print('total', total, 'mapped', mapped, 'uncat', method_counts['uncategorized'])
print('methods', dict(method_counts))
print('roots(count):', [(r['name_ar'], r['product_count_total']) for r in roots])
print('brands', len(brand_counts))
